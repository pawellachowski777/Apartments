#!/usr/bin/env python


def adverts(pages=1):
    """
    This function downloads tabular data about flats for rent in Warsaw from oxl.pl web side.
    """

    from urllib.request import urlopen 
    from bs4 import BeautifulSoup
    
    my_url = 'https://www.olx.pl/nieruchomosci/mieszkania/wynajem/warszawa/'
    links = []

    for i in range(pages+1):
        
        if pages == 0:
            break
            
        elif i >= 1:
            url = my_url + '?page=' + str(i)
            client = urlopen(url)
            page_html = client.read()
            page_soup = BeautifulSoup(page_html, 'html.parser')
            containers = page_soup.find_all('tr', {'class': 'wrap'})
            for cont in containers:
                link = cont.tr.td.a['href']
                if 'olx' in link:
                    links.append(link)
                    
    rows = []
    for link in list(set(links)):
        try:
            ad = urlopen(link).read()
            ad_soup = BeautifulSoup(ad, features="html.parser")
            
            added_class = ad_soup.find_all('li', {'class': 'offer-bottombar__item'})
            added = added_class[0].text.strip().split(', ')[1]
            
            localization_class = ad_soup.find_all('div', {'class': 'offer-user__address'})
            city = localization_class[0].text.split()[0]
            district = localization_class[0].text.split()[2]

            price_class = ad_soup.find_all('div', {'class': 'pricelabel'})
            price = price_class[0].text.strip().split('\n')[0]
            price = int(price.replace(' ', '').replace('zł', ''))

            by_class = ad_soup.find_all('ul', {'class': 'offer-details'})
            by = by_class[0].a.strong.text

            level_class = ad_soup.find_all('li', {'class': 'offer-details__item'})
            level = level_class[1].a.text.strip().split('\n')[1]

            furniture_class = ad_soup.find_all('li', {'class': 'offer-details__item'})
            furniture = furniture_class[2].a.text.strip().replace('\n', ' ').split()[1]
            if furniture == 'Tak':
                furniture = 1
            else:
                furniture = 0

            building_class = ad_soup.find_all('li', {'class': 'offer-details__item'})
            building = building_class[3].a.text.split()[2]

            surface_class = ad_soup.find_all('li', {'class': 'offer-details__item'})
            surface = int(surface_class[4].text.strip().replace('\n', ' ').split()[1])

            rooms_class = ad_soup.find_all('li', {'class': 'offer-details__item'})
            rooms = rooms_class[5].text.strip().replace('\n', ' ').split()[2]

            rent_class = ad_soup.find_all('li', {'class': 'offer-details__item'})
            rent = int(rent_class[6].text.strip().replace('\n', ' ').split()[2])
            
            rows.append([link, added, city, district, price, by, level, furniture,
                         building, surface, rooms, rent])
        except Exception:
            pass
    print('pobrano dane z', len(rows), 'ogłoszeń.')
    return rows


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook
    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def drop_value(df, column, value):
    """ funkcja do usuwania wybranych wartości """
    index = df[df[column] == value].index
    df.drop(index=index, inplace=True)
    df.reset_index(drop=True, inplace=True)
